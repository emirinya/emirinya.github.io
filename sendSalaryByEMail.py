#test environment win10_64 python_3.6.6

#coding:utf-8
import sys
import smtplib
from smtplib import SMTP
from email.mime.text import MIMEText
from email.header import Header

from openpyxl import load_workbook
import json

import time
import logging
import re

logger = logging.getLogger("test.conf")   #创建一个logger,默认为root logger
logger.setLevel(logging.DEBUG)   #设置全局log级别为debug。注意全局的优先级最高

hterm =  logging.StreamHandler()    #创建一个终端输出的handler,设置级别为error
hterm.setLevel(logging.INFO)

hfile = logging.FileHandler("access.log")    #创建一个文件记录日志的handler,设置级别为info
hfile.setLevel(logging.INFO)

formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')   #创建一个全局的日志格式

hterm.setFormatter(formatter)   #将日志格式应用到终端handler
hfile.setFormatter(formatter)   #将日志格式应用到文件handler

logger.addHandler(hterm)    #将终端handler添加到logger
logger.addHandler(hfile)    #将文件handler添加到logger

# 读取excel2007文件
wb = load_workbook(filename=r'salary.xlsx',data_only = True)
'''
# 显示有多少张表
print("Worksheet range(s):", wb.get_named_ranges())
print("Worksheet name(s):", wb.get_sheet_names())

# 取第一张表
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
'''

bolGetSenderInfo = True		# 发送者账号信息采集标志
# fmData = { 'Email Address' : [ 'Name' , ['Data in Table',...] ] }
fmData = {}
subject = ''
sender = ''
password = ''
helloword = ''
sheetNameList = wb.get_sheet_names()

for nub in sheetNameList:
	ws = wb.get_sheet_by_name(nub)
	
	if(bolGetSenderInfo):
		# -- Excel表格数据采集部分

		# 标题内容
		subject = ws['B2'].value + ws['D2'].value + '工资条'

		# 发送者的登陆用户名和密码
		sender = ws['F2'].value		# 发送者邮箱账号
		password = ws['H2'].value	# 发送者邮箱密码
		helloword = ws['J2'].value	# 问候语
		bolGetSenderInfo = False	

	for row in range(5,1000) :
		rowDataValue = []
		flagEMExist = False
		rowEMAddress = str( ws['AI'+str(row)].value )
		if re.match(r'^[0-9a-zA-Z_]{0,19}@[0-9a-zA-Z]{1,13}\.[com,cn,net]{1,3}$',rowEMAddress ):
			if rowEMAddress in fmData:
				flagEMExist = True
			else:
				rowName = str(ws['C'+str(row)].value)
		else:
			break
		
		dataValue = ''
		rowList = ws['B'+str(row):'W'+str(row)]
		for box in rowList[0] :
			dataValue += ('<td>'+str(box.value)+'</td>')
		
		if flagEMExist :
			fmData[rowEMAddress][1].append(dataValue)
		else:
			fmData[rowEMAddress] = [rowName,[dataValue]]

#	print(fmData,end='')

mailTextHeader1 = '<html><body><p>'
mailTextHeader2 = r' 你好：</p><p>' + str( helloword ) + '</p>'
mailTextTableHeader = '<table border="1"><tr><th rowspan="2">部门</th><th rowspan="2">姓名</th><th colspan="9">税前工资</th><th colspan="8">应扣部分</th><th rowspan="2">专项扣除</th><th rowspan="2">本月应缴个税</th><th rowspan="2">实发工资</th></tr><tr><th>小计</th><th>基本工资</th><th>绩效工资标准</th><th>绩效系数</th><th>实际绩效工资</th><th>其他补贴</th><th>加班费</th><th>奖金</th><th>请假扣款及其他扣款</th><th>小计</th><th>养老保险</th><th>门诊统筹基金</th><th>基本医疗保险</th><th>失业保险</th><th>公积金</th><th>公积金补扣</th><th>社保补扣</th></tr><tr>'
mailTextTableFooter = '</tr></table><p></p>'
mailTextFooter = '</body></html>'

# 发送者邮箱的SMTP服务器地址
smtpServer = 'smtp.yourmail.cn'
smtpServerPort = 465	# 默认端口是25 SSL端口为465 也可以根据服务器进行设定

# SMTP初始化
try:
	smtp = smtplib.SMTP_SSL(smtpServer,smtpServerPort) # 实例化SMTP对象
	smtp.login(sender,password) # 登陆smtp服务器
	logger.info(r'邮箱登录-成功')
except smtplib.SMTPException:
	logger.error(r'邮箱登录-失败')
	smtp.quit()
	sys.exit()
except BaseException:
	logger.error(r'SMTP初始化-失败')
	sys.exit()

countSec = 0
countFal = 0

for flag in fmData :
	mailTextTable = ''
	for tableData in fmData[flag][1] :
		mailTextTable += ( mailTextTableHeader + tableData + mailTextTableFooter )
	mailText = mailTextHeader1 + fmData[flag][0] + mailTextHeader2 + mailTextTable + mailTextFooter
	msg = MIMEText(mailText,'html','utf-8')
	msg['From'] = Header(sender,'utf-8')
	msg['To'] = Header(flag,'utf-8')
	msg['Subject'] = Header(subject,'utf-8') 
	
	print(mailText,end='')
	print("",end='')
	'''
	try:
		smtp.sendmail(sender,flag,msg.as_string())# 发送邮件 ，这里有三个参数
		time.sleep(3)
		logger.info(r'发送成功 - '+ flag)
		countSec += 1
	except smtplib.SMTPException as e:
		print(e)
		logger.error(r'发送失败 - '+ flag)
		countFal += 1
	except BaseException as e:
		print(e)
		logger.error(r'意外错误跳过发送'+ flag)
		logger.error(mailText)
		countFal += 1
	'''
smtp.quit()
logger.info(r'%d封邮件发送成功',countSec)
logger.info(r'%d封邮件发送失败',countFal)
